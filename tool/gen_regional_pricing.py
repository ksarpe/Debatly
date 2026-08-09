# -*- coding: utf-8 -*-
"""Cennik Debatly — ZAPIS STANU FAKTYCZNEGO z Google Play (nie propozycja).

Do 2026-08-09 ten plik generowal PROPOZYCJE cen z kotwic tierowych (US 8,99/24,99,
PL 24,99/69,99, lifetime ~2,8x miesiecznej). Ta propozycja zostala ODRZUCONA —
w Google Play stoi inny, docelowy cennik. Plik trzyma teraz TEN cennik, zeby
zaden pozniejszy "fix" nie cofnal sklepu do nieaktualnego planu.

Zrodlo: Google Play Console, ceny brutto (z VAT tam, gdzie kraj go nalicza),
stan na 2026-08-09. Poprzedni arkusz (exports/cennik_regionalny_2026-07-13.xlsx)
jest NIEAKTUALNY.

Arkusz liczy trzy rzeczy, ktorych w konsoli nie widac:
  * mnoznik lifetime/miesieczna i wynikajacy z niego podpis na paywallu
    ("Mniej niz N miesiecy subskrypcji" — patrz _lifetimeSubline w
    lib/features/monetization/widgets/paywall_offer_section.dart),
  * kwote, ktora realnie zostaje po VAT i 15% prowizji sklepu,
  * liste rynkow, gdzie mnoznik odstaje od reszty swiata.

Uruchomienie:  python tool/gen_regional_pricing.py
"""
import datetime
import math
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TODAY = "2026-08-09"

# Prowizja Google Play: 15% na pierwszy 1 mln USD rocznego przychodu oraz na
# subskrypcje od pierwszego dnia. Zmien, jesli przekroczysz prog.
STORE_FEE = 0.15

# Podpis na karcie lifetime liczy sie tak samo jak w apce: floor(L/M) + 1.
# Wartosc dominujaca w cenniku (patrz zakladka "Mnoznik") — odchylenia raportujemy.
TARGET_MONTHS = 4

# Waluty o innej niz 2 liczbie miejsc dziesietnych.
DECIMALS = {
    "JOD": 3,
    "JPY": 0, "KRW": 0, "CLP": 0, "COP": 0, "IDR": 0, "IQD": 0, "MMK": 0,
    "PKR": 0, "PYG": 0, "VND": 0, "XOF": 0, "XAF": 0, "HUF": 0, "RSD": 0,
}

# ---------------------------------------------------------------------------
# CENNIK: (kraj, waluta, miesieczna, dozywotnia, VAT % lub None, uwaga)
# Ceny dokladnie tak, jak stoja w Play Console (brutto).
# ---------------------------------------------------------------------------
PRICES = [
    ("Albania",                        "USD",  6.40,   22.08, 20.0,  ""),
    ("Algieria",                       "DZD",  479.00, 2450.00, None, ""),
    ("Angola",                         "USD",  5.33,   18.40, None,  ""),
    ("Antigua i Barbuda",              "USD",  5.33,   18.40, None,  ""),
    ("Arabia Saudyjska",               "SAR",  22.99,  78.99, 15.0,  ""),
    ("Argentyna",                      "USD",  5.33,   18.40, None,  ""),
    ("Armenia",                        "USD",  5.33,   18.40, None,  ""),
    ("Aruba",                          "USD",  5.33,   18.40, None,  ""),
    ("Australia",                      "AUD",  8.49,   31.99, 10.0,  ""),
    ("Austria",                        "EUR",  5.49,   18.99, 20.0,  ""),
    ("Azerbejdzan",                    "USD",  5.33,   18.40, None,  ""),
    ("Bahamy",                         "USD",  5.33,   18.40, None,  ""),
    ("Bahrajn",                        "USD",  5.99,   19.99, 10.0,  ""),
    ("Bangladesz",                     "BDT",  750.00, 2600.00, 15.0, ""),
    ("Belgia",                         "EUR",  5.49,   18.99, 21.0,  ""),
    ("Belize",                         "USD",  5.33,   18.40, None,  ""),
    ("Benin",                          "EUR",  4.66,   16.09, None,  ""),
    ("Bermudy",                        "USD",  5.49,   17.99, None,  ""),
    ("Bialorus",                       "USD",  6.40,   22.08, 20.0,  ""),
    ("Boliwia",                        "BOB",  36.99,  124.99, None, ""),
    ("Botswana",                       "USD",  5.33,   18.40, None,  ""),
    ("Bosnia i Hercegowina",           "USD",  5.33,   18.40, None,  ""),
    ("Brazylia",                       "BRL",  21.99,  59.99, None,  ""),
    ("Brytyjskie Wyspy Dziewicze",     "USD",  5.49,   17.99, None,  ""),
    ("Bulgaria",                       "EUR",  4.49,   18.99, 20.0,  "miesieczna nizsza niz reszta strefy EUR"),
    ("Burkina Faso",                   "EUR",  4.66,   16.09, None,  ""),
    ("Chile",                          "CLP",  4990,   20200, 19.0,  ""),
    ("Chorwacja",                      "EUR",  5.99,   19.99, 25.0,  ""),
    ("Côte d'Ivoire",                  "XOF",  3600,   12500, 18.0,  ""),
    ("Cypr",                           "EUR",  5.49,   18.99, 19.0,  ""),
    ("Czad",                           "USD",  5.33,   18.40, None,  ""),
    ("Czechy",                         "CZK",  139.99, 469.99, 21.0, ""),
    ("Dania",                          "DKK",  44.00,  159.00, 25.0, "+ korekty dla 2 terytoriow"),
    ("Demokratyczna Republika Konga",  "USD",  5.33,   18.40, None,  ""),
    ("Dominika",                       "USD",  5.33,   18.40, None,  ""),
    ("Dominikana",                     "USD",  5.33,   18.40, None,  ""),
    ("Dzibuti",                        "USD",  5.33,   18.40, None,  ""),
    ("Egipt",                          "EGP",  299.99, 1049.99, 14.0, ""),
    ("Ekwador",                        "USD",  5.49,   17.99, None,  ""),
    ("Erytrea",                        "USD",  5.33,   18.40, None,  ""),
    ("Estonia",                        "EUR",  5.99,   19.99, 24.0,  ""),
    ("Fidzi",                          "USD",  5.33,   18.40, None,  ""),
    ("Filipiny",                       "PHP",  365.00, 1250.00, 12.0, ""),
    ("Finlandia",                      "EUR",  5.99,   19.99, 25.5, "+ korekta dla 1 terytorium"),
    ("Francja",                        "EUR",  5.49,   18.99, 20.0, "+ korekty dla 11 terytoriow"),
    ("Gabon",                          "EUR",  4.66,   16.09, None,  ""),
    ("Gambia",                         "USD",  5.33,   18.40, None,  ""),
    ("Ghana",                          "GHS",  75.00,  250.00, 20.0, ""),
    ("Gibraltar",                      "GBP",  3.99,   13.99, None,  ""),
    ("Grecja",                         "EUR",  5.99,   19.99, 24.0,  ""),
    ("Grenada",                        "USD",  5.33,   18.40, None,  ""),
    ("Gruzja",                         "GEL",  19.00,  59.00, 18.0,  ""),
    ("Gwatemala",                      "USD",  5.33,   18.40, None,  ""),
    ("Gwinea",                         "USD",  5.49,   17.99, None,  ""),
    ("Gwinea Bissau",                  "EUR",  4.66,   16.09, None,  ""),
    ("Haiti",                          "USD",  5.33,   18.40, None,  ""),
    ("Hiszpania",                      "EUR",  5.49,   18.99, 21.0, "+ korekty dla 4 terytoriow"),
    ("Holandia",                       "EUR",  5.49,   18.99, 21.0,  ""),
    ("Honduras",                       "USD",  5.33,   18.40, None,  ""),
    ("Hongkong",                       "HKD",  43.00,  148.00, None, ""),
    ("Indie",                          "INR",  350.00, 849.00, 18.0, ""),
    ("Indonezja",                      "IDR",  96000,  329000, None, ""),
    ("Irak",                           "IQD",  7000,   24100, None,  ""),
    ("Irlandia",                       "EUR",  5.49,   19.99, 23.0,  ""),
    ("Islandia",                       "EUR",  5.78,   19.95, 24.0,  ""),
    ("Izrael",                         "ILS",  16.00,  55.00, None,  ""),
    ("Jamajka",                        "USD",  5.33,   18.40, None,  ""),
    ("Japonia",                        "JPY",  950,    3260,  10.0,  ""),
    ("Jemen",                          "USD",  5.33,   18.40, None,  ""),
    ("Jordania",                       "JOD",  3.800,  13.050, None, ""),
    ("Kajmany",                        "USD",  5.49,   17.99, None,  ""),
    ("Kambodza",                       "USD",  5.49,   17.99, None,  ""),
    ("Kamerun",                        "XAF",  3600,   12600, 19.25, ""),
    ("Kanada",                         "CAD",  7.49,   29.99, None,  ""),
    ("Katar",                          "QAR",  19.00,  67.00, None,  ""),
    ("Kazachstan",                     "KZT",  2890.00, 9990.00, 16.0, ""),
    ("Kenia",                          "KES",  800.00, 2800.00, 16.0, ""),
    ("Kirgistan",                      "USD",  5.33,   18.40, None,  ""),
    ("Kolumbia",                       "COP",  18000,  61000, None,  ""),
    ("Komory",                         "USD",  5.33,   18.40, None,  ""),
    ("Kongo",                          "USD",  5.33,   18.40, None,  ""),
    ("Korea Poludniowa",               "KRW",  9000,   31000, 10.0,  ""),
    ("Kostaryka",                      "CRC",  2400.00, 8300.00, None, ""),
    ("Kuwejt",                         "USD",  5.49,   17.99, None,  ""),
    ("Laos",                           "USD",  5.33,   18.40, None,  ""),
    ("Liban",                          "USD",  5.33,   18.40, None,  ""),
    ("Liberia",                        "USD",  5.33,   18.40, None,  ""),
    ("Libia",                          "USD",  5.33,   18.40, None,  ""),
    ("Liechtenstein",                  "CHF",  4.60,   16.00, 8.1,   ""),
    ("Litwa",                          "EUR",  5.49,   18.99, 21.0,  ""),
    ("Luksemburg",                     "EUR",  5.49,   18.99, 17.0,  ""),
    ("Lotwa",                          "EUR",  5.49,   18.99, 21.0,  ""),
    ("Macedonia Polnocna",             "USD",  5.33,   18.40, None,  ""),
    ("Makau",                          "MOP",  42.99,  149.00, None, ""),
    ("Malediwy",                       "USD",  5.33,   18.40, None,  ""),
    ("Malezja",                        "MYR",  21.99,  80.99, 8.0,   ""),
    ("Mali",                           "EUR",  4.66,   16.09, None,  ""),
    ("Malta",                          "EUR",  5.50,   18.99, 18.0,  ""),
    ("Maroko",                         "MAD",  59.99,  204.99, 20.0, ""),
    ("Mauritius",                      "USD",  5.33,   18.40, None,  ""),
    ("Meksyk",                         "MXN",  89.00,  375.00, 16.0, ""),
    ("Mikronezja",                     "USD",  5.49,   17.99, None,  ""),
    ("Mjanma (Birma)",                 "MMK",  11000,  39000, None,  ""),
    ("Moldawia",                       "USD",  6.40,   22.08, 20.0,  ""),
    ("Monako",                         "EUR",  5.49,   18.99, 20.0,  ""),
    ("Mongolia",                       "MNT",  19100.00, 65900.00, None, ""),
    ("Mozambik",                       "USD",  5.33,   18.40, None,  ""),
    ("Namibia",                        "USD",  5.33,   18.40, None,  ""),
    ("Nepal",                          "USD",  6.02,   20.79, 13.0,  ""),
    ("Niemcy",                         "EUR",  5.49,   18.99, 19.0,  ""),
    ("Niger",                          "EUR",  4.66,   16.09, None,  ""),
    ("Nigeria",                        "NGN",  7850.00, 27000.00, 7.5, ""),
    ("Nikaragua",                      "USD",  5.33,   18.40, None,  ""),
    ("Norwegia",                       "NOK",  65.00,  269.00, 25.0, "+ korekta dla 1 terytorium"),
    ("Nowa Zelandia",                  "NZD",  10.99,  36.99, 15.0,  ""),
    ("Oman",                           "USD",  5.49,   18.99, 5.0,   ""),
    ("Pakistan",                       "PKR",  1500,   5100,  None,  ""),
    ("Panama",                         "USD",  5.49,   17.99, None,  ""),
    ("Papua-Nowa Gwinea",              "USD",  5.33,   18.40, None,  ""),
    ("Paragwaj",                       "PYG",  30000,  100000, None, ""),
    ("Peru",                           "PEN",  17.99,  62.99, None,  ""),
    ("Polska",                         "PLN",  19.99,  69.99, 23.0,  "RYNEK BAZOWY"),
    ("Portugalia",                     "EUR",  5.49,   19.99, 23.0,  ""),
    ("Republika Poludniowej Afryki",   "ZAR",  74.99,  199.99, 15.0, ""),
    ("Republika Srodkowoafrykanska",   "EUR",  4.66,   16.09, None,  ""),
    ("Republika Zielonego Przyladka",  "USD",  5.33,   18.40, None,  ""),
    ("Rosja",                          "RUB",  409.00, 1399.00, None, ""),
    ("Rumunia",                        "RON",  22.99,  69.99, 21.0,  ""),
    ("Rwanda",                         "USD",  5.33,   18.40, None,  ""),
    ("Saint Kitts i Nevis",            "USD",  5.33,   18.40, None,  ""),
    ("Saint Lucia",                    "USD",  5.33,   18.40, None,  ""),
    ("Salwador",                       "USD",  5.49,   17.99, None,  ""),
    ("Samoa",                          "USD",  5.33,   18.40, None,  ""),
    ("San Marino",                     "EUR",  4.69,   15.99, None,  ""),
    ("Senegal",                        "XOF",  3600,   12500, 18.0,  ""),
    ("Serbia",                         "RSD",  649,    2299,  20.0,  ""),
    ("Seszele",                        "USD",  5.33,   18.40, None,  ""),
    ("Sierra Leone",                   "USD",  5.33,   18.40, None,  ""),
    ("Singapur",                       "SGD",  7.49,   25.98, 9.0,   ""),
    ("Slowacja",                       "EUR",  5.49,   19.99, 23.0,  ""),
    ("Slowenia",                       "EUR",  5.49,   19.99, 22.0,  ""),
    ("Somalia",                        "USD",  5.33,   18.40, None,  ""),
    ("Sri Lanka",                      "LKR",  1775.00, 6175.00, None, ""),
    ("Stany Zjednoczone",              "USD",  5.49,   22.99, None, "+ korekty dla 7 terytoriow"),
    ("Surinam",                        "USD",  5.33,   18.40, None,  ""),
    ("Szwajcaria",                     "CHF",  4.30,   19.00, None,  ""),
    ("Szwecja",                        "SEK",  65.00,  225.00, 25.0, ""),
    ("Tadzykistan",                    "USD",  5.33,   18.40, None,  ""),
    ("Tajlandia",                      "THB",  139.00, 379.00, 7.0,  ""),
    ("Tajwan",                         "TWD",  180.00, 620.00, 5.0,  ""),
    ("Tanzania",                       "TZS",  14000.00, 48000.00, None, ""),
    ("Togo",                           "EUR",  4.66,   16.09, None,  ""),
    ("Tonga",                          "USD",  5.33,   18.40, None,  ""),
    ("Trynidad i Tobago",              "USD",  5.33,   18.40, None,  ""),
    ("Tunezja",                        "USD",  5.33,   18.40, None,  ""),
    ("Turcja",                         "TRY",  299.99, 1029.99, 20.0, ""),
    ("Turkmenistan",                   "USD",  5.33,   18.40, None,  ""),
    ("Turks i Caicos",                 "USD",  5.49,   17.99, None,  ""),
    ("Uganda",                         "USD",  6.29,   21.71, 18.0,  ""),
    ("Ukraina",                        "UAH",  159.99, 979.99, 20.0, ""),
    ("Urugwaj",                        "USD",  5.33,   18.40, None,  ""),
    ("Uzbekistan",                     "USD",  5.97,   20.61, 12.0,  ""),
    ("Vanuatu",                        "USD",  5.33,   18.40, None,  ""),
    ("Watykan",                        "EUR",  4.69,   15.99, None,  ""),
    ("Wenezuela",                      "USD",  5.33,   18.40, None,  ""),
    ("Wielka Brytania",                "GBP",  4.79,   16.49, 20.0,  ""),
    ("Wietnam",                        "VND",  140000, 484000, None, ""),
    ("Wlochy",                         "EUR",  5.49,   19.99, 22.0,  ""),
    ("Wyspy Salomona",                 "USD",  5.33,   18.40, None,  ""),
    ("Wegry",                          "HUF",  1790,   4999,  27.0,  ""),
    ("Zambia",                         "USD",  5.33,   18.40, None,  ""),
    ("Zimbabwe",                       "USD",  5.33,   18.40, None,  ""),
    ("Zjednoczone Emiraty Arabskie",   "AED",  20.99,  70.99, 5.0,   ""),
    ("Pozostale / nowe kraje",         "USD",  5.33,   18.40, None, "domyslna cena Google dla nieobsadzonych rynkow"),
]

# ---------------------------------------------------------------------------
# Wyliczenia
# ---------------------------------------------------------------------------


def months_label(monthly, lifetime):
    """Ile pokaze paywall: floor(L/M) + 1, dokladnie jak _lifetimeSubline."""
    if monthly <= 0:
        return None
    return math.floor(lifetime / monthly) + 1


def net_of(price, vat):
    """Ile zostaje po VAT i prowizji sklepu."""
    base = price / (1 + vat / 100.0) if vat else price
    return base * (1 - STORE_FEE)


def fmt(value, currency):
    dec = DECIMALS.get(currency, 2)
    return f"{value:,.{dec}f}".replace(",", " ").replace(".", ",")


ROWS = []
for country, cur, monthly, lifetime, vat, note in PRICES:
    ratio = lifetime / monthly
    ROWS.append({
        "kraj": country,
        "waluta": cur,
        "m": monthly,
        "l": lifetime,
        "vat": vat,
        "ratio": ratio,
        "months": months_label(monthly, lifetime),
        "m_net": net_of(monthly, vat),
        "l_net": net_of(lifetime, vat),
        "uwaga": note,
    })

# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
GOOD_FILL = PatternFill("solid", fgColor="DCFCE7")
BASE_FILL = PatternFill("solid", fgColor="DBEAFE")
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_header(ws, headers, widths):
    for i, (title, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=i, value=title)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def sheet_cennik(wb):
    ws = wb.create_sheet("Cennik")
    write_header(
        ws,
        ["Kraj", "Waluta", "Miesieczna", "Dozywotnia", "Mnoznik L/M",
         "Paywall pokaze", "VAT", "Netto / mies.", "Netto / lifetime", "Uwaga"],
        [30, 8, 14, 14, 12, 16, 8, 14, 16, 42],
    )
    for r, row in enumerate(ROWS, start=2):
        ws.cell(row=r, column=1, value=row["kraj"])
        ws.cell(row=r, column=2, value=row["waluta"])
        ws.cell(row=r, column=3, value=fmt(row["m"], row["waluta"]))
        ws.cell(row=r, column=4, value=fmt(row["l"], row["waluta"]))
        ws.cell(row=r, column=5, value=round(row["ratio"], 2))
        ws.cell(row=r, column=6, value=f'mniej niz {row["months"]} mies.')
        ws.cell(row=r, column=7, value="—" if row["vat"] is None else f'{row["vat"]}%')
        ws.cell(row=r, column=8, value=fmt(row["m_net"], row["waluta"]))
        ws.cell(row=r, column=9, value=fmt(row["l_net"], row["waluta"]))
        ws.cell(row=r, column=10, value=row["uwaga"])
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = BORDER
            if c >= 3:
                ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
        if row["kraj"] == "Polska":
            for c in range(1, 11):
                ws.cell(row=r, column=c).fill = BASE_FILL
        elif row["months"] != TARGET_MONTHS:
            ws.cell(row=r, column=6).fill = (
                WARN_FILL if row["months"] > TARGET_MONTHS else GOOD_FILL
            )
    ws.auto_filter.ref = f"A1:J{len(ROWS) + 1}"


def sheet_mnoznik(wb):
    ws = wb.create_sheet("Mnoznik")
    write_header(
        ws,
        ["Kraj", "Waluta", "Miesieczna", "Dozywotnia", "Mnoznik", "Paywall pokaze",
         "Ocena", "Min. miesieczna dla 'mniej niz 4'"],
        [30, 8, 14, 14, 10, 16, 34, 30],
    )
    odd = [r for r in ROWS if r["months"] != TARGET_MONTHS]
    odd.sort(key=lambda r: (-r["months"], r["kraj"]))
    for r, row in enumerate(odd, start=2):
        if row["months"] > TARGET_MONTHS:
            verdict = "slabszy anchor — lifetime drogi wzgledem miesiecznej"
            need = fmt(row["l"] / TARGET_MONTHS, row["waluta"]) + " (powyzej tej kwoty)"
        else:
            verdict = "mocniejszy anchor — zostawic"
            need = "—"
        ws.cell(row=r, column=1, value=row["kraj"])
        ws.cell(row=r, column=2, value=row["waluta"])
        ws.cell(row=r, column=3, value=fmt(row["m"], row["waluta"]))
        ws.cell(row=r, column=4, value=fmt(row["l"], row["waluta"]))
        ws.cell(row=r, column=5, value=round(row["ratio"], 2))
        ws.cell(row=r, column=6, value=f'mniej niz {row["months"]} mies.')
        ws.cell(row=r, column=7, value=verdict)
        ws.cell(row=r, column=8, value=need)
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).fill = (
                WARN_FILL if row["months"] > TARGET_MONTHS else GOOD_FILL
            )
    note = ws.cell(row=len(odd) + 3, column=1)
    note.value = (
        f"Reszta swiata ({len(ROWS) - len(odd)} z {len(ROWS)} rynkow) pokazuje "
        f'"mniej niz {TARGET_MONTHS} mies." — to jest norma, wzgledem ktorej '
        "liczona jest ta lista."
    )
    note.font = Font(italic=True, size=9)


def sheet_ios(wb):
    ws = wb.create_sheet("iOS do wpisania")
    write_header(
        ws,
        ["Kraj", "Waluta", "Miesieczna (cel)", "Dozywotnia (cel)", "Wpisane w ASC?"],
        [30, 8, 18, 18, 16],
    )
    for r, row in enumerate(ROWS, start=2):
        ws.cell(row=r, column=1, value=row["kraj"])
        ws.cell(row=r, column=2, value=row["waluta"])
        ws.cell(row=r, column=3, value=fmt(row["m"], row["waluta"]))
        ws.cell(row=r, column=4, value=fmt(row["l"], row["waluta"]))
        ws.cell(row=r, column=5, value="")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = BORDER
    ws.auto_filter.ref = f"A1:E{len(ROWS) + 1}"


def sheet_info(wb):
    ws = wb.create_sheet("Info", 0)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 118
    lines = [
        ("H", f"Cennik Debatly — stan Google Play na {TODAY}"),
        ("", ""),
        ("H", "Co to jest"),
        ("", "Zapis CEN, KTORE STOJA W SKLEPIE — nie propozycja. Google Play jest zrodlem prawdy."),
        ("", "Ceny sa brutto (z VAT tam, gdzie kraj go nalicza), dokladnie jak w Play Console."),
        ("", "Arkusz exports/cennik_regionalny_2026-07-13.xlsx (propozycja tierowa 8,99/24,99 USD,"),
        ("", "PL 24,99/69,99) jest NIEAKTUALNY — zostal odrzucony, nie wracaj do niego."),
        ("", ""),
        ("H", "Zakladki"),
        ("", "Cennik — pelna lista + mnoznik, podpis paywalla i kwota netto po VAT i prowizji."),
        ("", "Mnoznik — rynki, gdzie podpis odstaje od reszty swiata (i o ile trzeba ruszyc cene)."),
        ("", "iOS do wpisania — ta sama lista jako checklista dla App Store Connect."),
        ("", ""),
        ("H", "Dlaczego mnoznik ma znaczenie"),
        ("", "Karta lifetime na paywallu pokazuje 'Mniej niz N miesiecy subskrypcji', gdzie"),
        ("", "N = floor(lifetime / miesieczna) + 1 — liczone w apce z cen zwroconych przez"),
        ("", "RevenueCat, wiec zmiana ceny w sklepie NATYCHMIAST zmienia ten podpis."),
        ("", "Kod: _lifetimeSubline w lib/features/monetization/widgets/paywall_offer_section.dart."),
        ("", "Nizsze N = mocniejszy argument za lifetime. Podpis znika przy N < 2."),
        ("", ""),
        ("H", "Netto"),
        ("", f"cena / (1 + VAT) x (1 - {STORE_FEE:.0%} prowizji sklepu)."),
        ("", "15% to stawka Google na pierwszy 1 mln USD rocznie oraz na subskrypcje od 1. dnia."),
        ("", "Podatek dochodowy NIE jest tu uwzgledniony."),
        ("", ""),
        ("H", "Aktualizacja"),
        ("", "Zmien tablice PRICES w tool/gen_regional_pricing.py i uruchom:"),
        ("", "    python tool/gen_regional_pricing.py"),
    ]
    for i, (kind, text) in enumerate(lines, start=2):
        cell = ws.cell(row=i, column=2, value=text)
        if kind == "H":
            cell.font = Font(bold=True, size=12 if i == 2 else 11, color="1F2937")
        else:
            cell.alignment = Alignment(wrap_text=False)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    sheet_cennik(wb)
    sheet_mnoznik(wb)
    sheet_ios(wb)
    sheet_info(wb)

    out_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "exports")
    os.makedirs(out_dir, exist_ok=True)
    path = os.path.join(out_dir, f"cennik_google_play_{TODAY}.xlsx")
    wb.save(path)

    odd = [r for r in ROWS if r["months"] != TARGET_MONTHS]
    print(f"Zapisano: {path}")
    print(f"Rynkow: {len(ROWS)}  |  odstajacy mnoznik: {len(odd)}")
    for row in sorted(odd, key=lambda r: -r["ratio"]):
        print(
            f'  {row["kraj"]:<32} {row["waluta"]}  '
            f'{fmt(row["m"], row["waluta"])} / {fmt(row["l"], row["waluta"])}  '
            f'= {row["ratio"]:.2f}x  -> "mniej niz {row["months"]} mies."'
        )


if __name__ == "__main__":
    main()
